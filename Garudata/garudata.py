import os
from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
import json
import csv
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['JSON_FOLDER'] = 'json_files'
app.config['DB_FILE'] = 'data.db'
app.config['ALLOWED_EXTENSIONS'] = {'csv', 'xlsx', 'db'}

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

if not os.path.exists(app.config['JSON_FOLDER']):
    os.makedirs(app.config['JSON_FOLDER'])

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def db_to_json(db_file, table_name):
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {table_name}")
        columns = [column[0] for column in cursor.description]
        data = cursor.fetchall()
        json_data = []
        for row in data:
            json_data.append(dict(zip(columns, row)))
        return json_data
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn:
            conn.close()

def csv_to_json(csv_file):
    try:
        json_data = []
        with open(csv_file, 'r') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                json_data.append(row)
        return json_data
    except Exception as e:
        return {"error": str(e)}

def xlsx_to_json(xlsx_file):
    try:
        json_data = []
        workbook = load_workbook(filename=xlsx_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            json_data.append(dict(zip(headers, row)))
        return json_data
    except Exception as e:
        return {"error": str(e)}
    
def read_json_files(directory):
    json_data = []
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            with open(os.path.join(directory, filename), 'r') as file:
                json_data.extend(json.load(file))
    return json_data

def search_json(json_data, keywords):
    results = []
    for item in json_data:
        for key, value in item.items():
            if keywords.lower() in str(value).lower():
                results.append(item)
                break
    return results

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert/db', methods=['POST'])
def convert_db_to_json():
    if 'db_file' not in request.files:
        return jsonify({"error": "No file part in the request."}), 400
    file = request.files['db_file']
    if file.filename == '':
        return jsonify({"error": "No file selected for uploading."}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        data = request.form
        table_name = data.get('table_name')
        if not table_name:
            return jsonify({"error": "Table name is required."}), 400

        json_data = db_to_json(file_path, table_name)
        if "error" in json_data:
            return jsonify(json_data), 400

        json_filename = f"{table_name}.json"
        json_path = os.path.join(app.config['JSON_FOLDER'], json_filename)
        with open(json_path, 'w') as json_file:
            json.dump(json_data, json_file)

        try:
            conn = sqlite3.connect(app.config['DB_FILE'])
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS json_data (
                    id INTEGER PRIMARY KEY,
                    file_name TEXT,
                    table_name TEXT
                )
            ''')
            cursor.execute('''
                INSERT INTO json_data (file_name, table_name)
                VALUES (?, ?)
            ''', (filename, table_name))
            conn.commit()
            return jsonify({"success": True, "message": "JSON file created successfully.", "json_file_path": json_path})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        finally:
            if conn:
                conn.close()

    return jsonify({"error": "Allowed file types are csv, xlsx, db."}), 400

@app.route('/convert/csv', methods=['POST'])
def convert_csv_to_json():
    if 'csv_file' not in request.files:
        return jsonify({"error": "No file part in the request."}), 400
    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({"error": "No file selected for uploading."}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        json_data = csv_to_json(file_path)
        if "error" in json_data:
            return jsonify(json_data), 400

        json_filename = f"{os.path.splitext(filename)[0]}.json"
        json_path = os.path.join(app.config['JSON_FOLDER'], json_filename)
        with open(json_path, 'w') as json_file:
            json.dump(json_data, json_file)

        try:
            conn = sqlite3.connect(app.config['DB_FILE'])
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS json_data (
                    id INTEGER PRIMARY KEY,
                    file_name TEXT,
                    table_name TEXT
                )
            ''')
            cursor.execute('''
                INSERT INTO json_data (file_name, table_name)
                VALUES (?, ?)
            ''', (filename, ''))
            conn.commit()
            return jsonify({"success": True, "message": "JSON file created successfully.", "json_file_path": json_path})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        finally:
            if conn:
                conn.close()

    return jsonify({"error": "Allowed file types are csv, xlsx, db."}), 400

@app.route('/convert/xlsx', methods=['POST'])
def convert_xlsx_to_json():
    if 'xlsx_file' not in request.files:
        return jsonify({"error": "No file part in the request."}), 400
    file = request.files['xlsx_file']
    if file.filename == '':
        return jsonify({"error": "No file selected for uploading."}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        json_data = xlsx_to_json(file_path)
        if "error" in json_data:
            return jsonify(json_data), 400

        json_filename = f"{os.path.splitext(filename)[0]}.json"
        json_path = os.path.join(app.config['JSON_FOLDER'], json_filename)
        with open(json_path, 'w') as json_file:
            json.dump(json_data, json_file)

        try:
            conn = sqlite3.connect(app.config['DB_FILE'])
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS json_data (
                    id INTEGER PRIMARY KEY,
                    file_name TEXT,
                    table_name TEXT
                )
            ''')
            cursor.execute('''
                INSERT INTO json_data (file_name, table_name)
                VALUES (?, ?)
            ''', (filename, ''))
            conn.commit()
            return jsonify({"success": True, "message": "JSON file created successfully.", "json_file_path": json_path})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        finally:
            if conn:
                conn.close()

    return jsonify({"error": "Allowed file types are csv, xlsx, db."}), 400

@app.route('/search', methods=['POST'])
def search():
    data = request.json
    directory = os.getcwd() + '/json_files'
    keywords = data.get('keywords')
    json_data = read_json_files(directory)
    results = search_json(json_data, keywords)
    
    if results:
        return jsonify(results)
    else:
        return jsonify({"error": "No results found."}), 400
    
@app.route('/get/json_data', methods=['GET'])
def get_json_data():
    try:
        conn = sqlite3.connect(app.config['DB_FILE'])
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM json_data ORDER BY id DESC LIMIT 10
        ''')
        rows = cursor.fetchall()
        json_list = []
        for row in rows:
            json_list.append({
                "file_name": row[1],
                "table_name": row[2]
            })
        return jsonify(json_list)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/download/json', methods=['GET'])
def download_json():
    file_name = request.args.get('file_name')
    if not file_name:
        return jsonify({"error": "file_name is required."}), 400

    json_path = os.path.join(app.config['JSON_FOLDER'], file_name)
    if not os.path.exists(json_path):
        return jsonify({"error": "File not found."}), 404

    return send_file(json_path, as_attachment=True)

@app.route('/get/preview', methods=['GET'])
def get_preview():
    file_name = request.args.get('file_name')
    if not file_name:
        return jsonify({"error": "file_name is required."}), 400

    json_path = os.path.join(app.config['JSON_FOLDER'], file_name)
    if not os.path.exists(json_path):
        return jsonify({"error": "File not found."}), 404

    try:
        with open(json_path, 'r') as json_file:
            data = json.load(json_file)
            first_10 = data[:10]
            return jsonify(first_10)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    if not os.path.exists(app.config['JSON_FOLDER']):
        os.makedirs(app.config['JSON_FOLDER'])
    app.run(debug=True)